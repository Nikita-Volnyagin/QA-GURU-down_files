import zipfile
from zipfile import ZipFile

from pypdf import PdfReader
from script_os import ARCHIVE_DIR, FILES_LIST
from io import TextIOWrapper
from openpyxl import load_workbook
import os
import csv
import download


def test_download_files_in_tmp():
    download.test_download_pdf_file()
    download.test_download_csv_file()
    download.test_download_xlsx_file()


def test_create_archive(create_archive):
    with ZipFile(ARCHIVE_DIR, mode='a') as zf:
        num_files = len(zf.filelist)
        list_files = []
        for file in zf.infolist():
            name = os.path.basename(file.filename)
            list_files.append(name)
    assert num_files==3
    assert FILES_LIST==list_files


def test_text_in_csv_file(create_archive):
    with zipfile.ZipFile(ARCHIVE_DIR) as zip_file:
        with zip_file.open('file_csv.csv') as csv_file:
            csvreader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8-sig')))
            assert 'White'==csvreader[1][0]


def test_text_in_pdf_file(create_archive):
    with zipfile.ZipFile(ARCHIVE_DIR) as zip_file:
        with zip_file.open('file_pdf.pdf') as pdf:
            reader = PdfReader(pdf)
            text = reader.pages[0].extract_text()
            assert 'pytest Documentation' in text


def test_text_in_xlsx_file(create_archive):
    with zipfile.ZipFile(ARCHIVE_DIR) as zip_file:
        with zip_file.open('file_xlsx.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            first_title = sheet.cell(row=1, column=1).value
            second_title = sheet.cell(row=1, column=2).value
            third_title = sheet.cell(row=1, column=3).value
            assert first_title=="Внешний идентификатор для импорта"
            assert second_title=="Вышестоящий отдел"
            assert third_title=="Название"